import openpyxl
clip = '1'

Arousal = [1,2,3,4,5]
Valence = [3,7,9,10,12]
def storeData():
    book = openpyxl.load_workbook('emotionData.xlsx')
    sheetname = 'Clip' + clip
    sheet = book[sheetname]
    col_count = sheet.max_column
    cur_col = col_count + 1

    if(col_count > 1):
        ratorNum = col_count/2 + 1
    else:
        ratorNum = 1

    sheet.cell(row = 1, column = cur_col).value = 'Rator ' + str(ratorNum) + ' (arousal)'
    sheet.cell(row = 1, column = cur_col + 1).value = 'Rator ' + str(ratorNum) + ' (valence)'

    j = 2
    for cell in Arousal:
        sheet.cell(row = j, column = cur_col).value = cell
        j = j + 1

    j = 2
    for cell in Valence:
        sheet.cell(row = j , column = cur_col + 1).value = cell
        j = j + 1



    book.save('emotionData.xlsx')